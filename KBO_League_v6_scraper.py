import requests
import urllib
import csv
import pandas as pd
import os
from bs4 import BeautifulSoup

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = writer.sheet_name.max_row
        # startrow=writer.sheets['Pitchers'].max_row,
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

os.chdir('C:\\Users\\Chris.Duke\\Desktop\\Temp Files\\Baseball\\KBO')
print (os.getcwd())

# Set the dates for which data will be scraped/updated
date_today = 10.24
week_today = 26


# opens the CSV list of players as "rosters" and reads it in.
with open('C:\\Users\\Chris.Duke\\Desktop\\Temp Files\\Baseball\\KBO\\KBO_League_v6_Player_List.csv') as rosters:
     csv_input = csv.DictReader(rosters)
     # create the empty data frames to which each player's data will be appended - one for pitchers and one for hitters
     pitcher_columns = ['Fantasy_Team','Fantasy_Position','KBO_URL','KBO_Player_Name','KBO_Team','KBO_Position','Date', 'Opponent', 'ERA', 'Result', 'PA', 'IP', 'H', 'HR', 'BB', 'HBP', 'K', 'R', 'ER', 'OAVG']
     todaystats_pitchers = pd.DataFrame(columns = pitcher_columns)
     hitter_columns= ['Fantasy_Team','Fantasy_Position','KBO_URL','KBO_Player_Name','KBO_Team','KBO_Position','Date','Opponent', 'AVG', 'AB', 'R', 'H', '2B', '3B', 'HR','RBI','SB','CS','BB','HBP','SO','GIDP','SF','TB','HQS']
     todaystats_hitters = pd.DataFrame(columns = hitter_columns)
     playercounter = 0
     # start iterating through the list of players - reading each row and putting together the data frame
     for row in csv_input:
         playercounter = playercounter + 1
         # Each player record in the rosters includes these data; store each in a variable so that it can be added to the data 
         Fantasy_Team = row['Fantasy_Team']
         Fantasy_Position = row['Fantasy_Position']
         KBO_URL = row['KBO_URL'] # This URL is the Player's Game Log at the KBO Official website
         KBO_Player_Name = row['KBO_Player_Name']
         KBO_Team = row['KBO_Team']
         KBO_Position = row['BR_Position']         
 
         # Determine if player is a pitcher - if so, collect that data and store it. If not, skip to the hitter section below.
         if KBO_Position == '1':
            # Create dfpitcher dataframe and read into it the last row of HTML data from the Player's Game Log URL
            dfpitcher = pd.read_html(KBO_URL)[-1]
            # Define the column names for the data read from the URL
            dfpitcher.columns = ['Date', 'Opponent', 'ERA', 'Result', 'PA', 'IP', 'H', 'HR', 'BB', 'HBP', 'K', 'R', 'ER', 'OAVG']
            dfpitcher['ERA'].astype('float')
            # Add the variables from the roster list to the data being scraped.
            dfpitcher['Fantasy_Team'] = Fantasy_Team
            dfpitcher['Fantasy_Position'] = Fantasy_Position
            dfpitcher['KBO_URL'] = KBO_URL.rstrip('\n')
            dfpitcher['KBO_Player_Name'] = KBO_Player_Name
            dfpitcher['KBO_Team'] = KBO_Team
            dfpitcher['KBO_Position'] = KBO_Position
            dfpitcher['Week'] = week_today
            # Check "Date" scraped from the last row of the player game log; only store that row of data if date = date_today
            todaystats = dfpitcher.loc[dfpitcher['Date'] == date_today ]
            # This is just a counter that runs in the console - that lets me know/see the progress.
            if todaystats.empty:
               print(playercounter,KBO_Player_Name,KBO_Team,'| NO STATS at',Fantasy_Position,'for',Fantasy_Team)
               continue
            print(playercounter,KBO_Player_Name,KBO_Team,'| played today at',Fantasy_Position,'for',Fantasy_Team)
            # These two rows put the columns in the same order that I have them in the Excel spreadsheet.
            reorder = ['Fantasy_Team','Fantasy_Position','KBO_URL','KBO_Player_Name','KBO_Team','KBO_Position','Week','Date', 'Opponent', 'ERA', 'Result', 'PA', 'IP', 'H', 'HR', 'BB', 'HBP', 'K', 'R', 'ER', 'OAVG']
            todaystats_pitchers = todaystats_pitchers.reindex(columns=reorder)
            # This adds the player's data to the temporary dataframe. Once all players have been checked, this dataframe is written to Excel.
            todaystats_pitchers = todaystats_pitchers.append(todaystats)


         if KBO_Position != '1':
            # Create dfhitter dataframe and read into it the last row of HTML data from the Player's Game Log URL
            dfhitter = pd.read_html(KBO_URL)[-1] 
            # Define the column names for the data read from the URL
            dfhitter.columns = ['Date', 'Opponent', 'AVG', 'AB', 'R', 'H', '2B', '3B', 'HR','RBI','SB','CS','BB','HBP','SO','GIDP']
            # Add the variables from the roster list to the data being scraped.
            dfhitter['Fantasy_Team'] = Fantasy_Team
            dfhitter['Fantasy_Position'] = Fantasy_Position
            dfhitter['KBO_URL'] = KBO_URL.rstrip('\n')
            dfhitter['KBO_Player_Name'] = KBO_Player_Name
            dfhitter['KBO_Team'] = KBO_Team
            dfhitter['KBO_Position'] = KBO_Position
            dfhitter['Week'] = week_today
            dfhitter['TB'] = ((dfhitter['H'] - dfhitter['2B'] - dfhitter['3B'] - dfhitter['HR']) + (2 * dfhitter['2B']) + (3 * dfhitter['3B']) + (4 * dfhitter['HR']))
            dfhitter['HQS'] = (dfhitter['R'] + dfhitter['H'] + dfhitter['HR'] + dfhitter['RBI'] + dfhitter['SB'] + dfhitter['BB'])
            # Check "Date" scraped from the last row of the player game log; only store that row of data if date = date_today
            todaystats = dfhitter.loc[dfhitter['Date'] == date_today ]
            # This is just a counter that runs in the console - that lets me know/see the progress.
            if todaystats.empty:
               print(playercounter,KBO_Player_Name,KBO_Team,'| NO STATS at',Fantasy_Position,'for',Fantasy_Team)
               continue
            print(playercounter,KBO_Player_Name,KBO_Team,'| played today at',Fantasy_Position,'for',Fantasy_Team)            # These two rwos put the columns in the same order that I have them in the Excel spreadsheet.
            # These two rows put the columns in the same order that I have them in the Excel spreadsheet.
            reorder = ['Fantasy_Team','Fantasy_Position','KBO_URL','KBO_Player_Name','KBO_Team','KBO_Position','Week','Date','Opponent', 'AVG', 'AB', 'R', 'H', '2B', '3B', 'HR','RBI','SB','CS','BB','HBP','SO','GIDP','SF','TB','HQS']
            todaystats_hitters = todaystats_hitters.reindex(columns=reorder)
            # This adds the player's data to the temporary dataframe. Once all players have been checked, this dataframe is written to Excel.
            todaystats_hitters = todaystats_hitters.append(todaystats)

# After all players have been checked - no more rows in Rosters CSV - write the data collection to the Excel file.
     print('Writing Pitcher Stats to Excel')
     append_df_to_excel('KBO_League_v6.xlsx', todaystats_pitchers, sheet_name='Pitchers', index=False, header=False, truncate_sheet=False)
     print('Writing Hitter Stats to Excel')
     append_df_to_excel('KBO_League_v6.xlsx', todaystats_hitters, sheet_name='Hitters', index=False, header=False, truncate_sheet=False)

# Close the rosters file. 
rosters.close()



